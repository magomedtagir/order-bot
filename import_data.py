"""
import_data.py — импорт товаров и клиентов из Excel в SQLite БД бота
Использование: python import_data.py order_bot_template_v2.xlsx [путь_к_бд]
"""

import sys
import sqlite3
import re
from openpyxl import load_workbook

EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else "order_bot_template_v2.xlsx"
DB_PATH = sys.argv[2] if len(sys.argv) > 2 else "orders.db"

def clean(val):
    if val is None:
        return ""
    return str(val).strip()

def init_tables(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY,
            full_name TEXT NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS product_aliases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL REFERENCES products(id),
            alias TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS client_aliases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER NOT NULL REFERENCES clients(id),
            alias TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS client_name_cache (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER NOT NULL REFERENCES clients(id),
            base_name TEXT NOT NULL,
            product_id INTEGER REFERENCES products(id),
            resolved_by TEXT,
            confidence REAL DEFAULT 1.0,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(client_id, base_name)
        );
    """)
    conn.commit()

def import_products(ws, conn):
    # Читаем все товары из Excel
    excel_products: dict[int, str] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        pid, full_name = row[0], clean(row[1])
        if not pid or not full_name:
            continue
        try:
            pid = int(pid)
        except (ValueError, TypeError):
            continue
        excel_products[pid] = full_name

    # Вставляем / обновляем
    inserted = updated = 0
    for pid, full_name in excel_products.items():
        existing = conn.execute("SELECT full_name FROM products WHERE id = ?", (pid,)).fetchone()
        if existing is None:
            conn.execute("INSERT INTO products (id, full_name) VALUES (?, ?)", (pid, full_name))
            inserted += 1
        elif existing[0] != full_name:
            conn.execute("UPDATE products SET full_name = ? WHERE id = ?", (full_name, pid))
            updated += 1

    # Удаляем позиции, которых больше нет в Excel
    db_ids = {row[0] for row in conn.execute("SELECT id FROM products").fetchall()}
    removed_ids = db_ids - set(excel_products.keys())
    deleted = 0
    for pid in removed_ids:
        conn.execute("DELETE FROM product_aliases WHERE product_id = ?", (pid,))
        conn.execute("DELETE FROM products WHERE id = ?", (pid,))
        deleted += 1

    conn.commit()
    print(f"  ✅ Товары: добавлено {inserted}, обновлено {updated}, удалено {deleted}")

def import_clients(ws, conn):
    inserted = 0
    alias_count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[1]:
            continue
        full_name = clean(row[1])
        aliases_raw = clean(row[2]) if len(row) > 2 else ""

        if not full_name:
            continue

        # Вставить клиента
        conn.execute("INSERT OR IGNORE INTO clients (full_name) VALUES (?)", (full_name,))
        client_id = conn.execute("SELECT id FROM clients WHERE full_name = ?", (full_name,)).fetchone()[0]
        if conn.execute("SELECT changes()").fetchone()[0]:
            inserted += 1

        # Вставить псевдонимы
        aliases = [a.strip() for a in aliases_raw.split(",") if a.strip()]
        # Добавить само полное название как псевдоним (в нижнем регистре)
        aliases.append(full_name.lower())
        for alias in set(aliases):
            alias_lower = alias.lower()
            try:
                conn.execute(
                    "INSERT OR IGNORE INTO client_aliases (client_id, alias) VALUES (?, ?)",
                    (client_id, alias_lower)
                )
                if conn.execute("SELECT changes()").fetchone()[0]:
                    alias_count += 1
            except Exception:
                pass

    conn.commit()
    print(f"  ✅ Клиенты: добавлено {inserted}, псевдонимов: {alias_count}")

def main():
    print(f"\n📂 Читаем файл: {EXCEL_PATH}")
    print(f"💾 База данных: {DB_PATH}\n")

    wb = load_workbook(EXCEL_PATH, read_only=True)
    print(f"Листы в файле: {wb.sheetnames}\n")

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")

    print("🔧 Создаём таблицы если нет...")
    init_tables(conn)

    # Товары
    if "Товары" in wb.sheetnames:
        print("📦 Импортируем товары...")
        import_products(wb["Товары"], conn)
    else:
        print("⚠️  Лист 'Товары' не найден, пропускаем")

    # Клиенты
    if "Клиенты" in wb.sheetnames:
        print("👥 Импортируем клиентов...")
        import_clients(wb["Клиенты"], conn)
    else:
        print("⚠️  Лист 'Клиенты' не найден, пропускаем")

    # Итог
    prod_count = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    client_count = conn.execute("SELECT COUNT(*) FROM clients").fetchone()[0]
    alias_count = conn.execute("SELECT COUNT(*) FROM client_aliases").fetchone()[0]

    print(f"\n📊 Итог в БД:")
    print(f"   Товаров:           {prod_count}")
    print(f"   Клиентов:          {client_count}")
    print(f"   Псевдонимов клиентов: {alias_count}")
    print(f"\n✅ Готово! Скопируй {DB_PATH} в папку бота.")

    conn.close()

if __name__ == "__main__":
    main()
